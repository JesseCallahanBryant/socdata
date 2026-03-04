"""wvs.py — World Values Survey Time Series provider (1981–2022)."""

from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd

from socdata.datasets.base import DatasetProvider, VariableInfo
from socdata.datasets.cache import (
    CACHE_DIR,
    cache_path,
    has_cache,
    load_cache,
    save_cache,
)
from socdata.datasets.registry import register

WVS_CACHE_KEY = "wvs_timeseries"
WVS_META_PATH = CACHE_DIR / "wvs_timeseries_meta.json"

# Where the user should place their downloaded files
WVS_DATA_DIR = CACHE_DIR / "World Values Survey"
WVS_CSV_NAME = "WVS_Time_Series_1981-2022_csv_v5_0.csv"
WVS_EXCEL_GLOB = "F00003844-WVS_Time_Series_List_of_Variables_and_equivalences*.xlsx"
WVS_DOWNLOAD_URL = "https://www.worldvaluessurvey.org/WVSDocumentationWVL.jsp"


def _load_meta() -> dict[str, str]:
    """Load variable labels from cached JSON."""
    if WVS_META_PATH.exists():
        meta = json.loads(WVS_META_PATH.read_text(encoding="utf-8"))
        return meta.get("variable_labels", {})
    return {}


def _save_meta(variable_labels: dict[str, str]) -> None:
    """Save metadata as JSON alongside the parquet cache."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    WVS_META_PATH.write_text(
        json.dumps({"variable_labels": variable_labels}, ensure_ascii=False),
        encoding="utf-8",
    )


def _load_variable_labels() -> dict[str, str]:
    """Read variable labels from the WVS documentation Excel file.

    The Excel file has: column B = variable name, column C = label/title.
    """
    excel_files = sorted(WVS_DATA_DIR.glob(WVS_EXCEL_GLOB))
    if not excel_files:
        return {}

    import openpyxl

    wb = openpyxl.load_workbook(str(excel_files[0]), read_only=True)
    ws = wb[wb.sheetnames[0]]
    labels: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
        var_name, title = row
        if var_name and title:
            labels[str(var_name).strip().upper()] = str(title).strip()
    wb.close()
    return labels


def _recode_missing(df: pd.DataFrame) -> pd.DataFrame:
    """Replace WVS negative missing-data codes with NaN.

    WVS convention: -1 = don't know, -2 = no answer,
    -4 = not asked in survey, -5 = missing/unknown.
    """
    numeric = df.select_dtypes(include="number")
    df[numeric.columns] = numeric.where(numeric >= 0, other=np.nan)
    return df


class WVSProvider(DatasetProvider):
    name = "wvs"
    display_name = "World Values Survey"
    description = "WVS Time Series (1981–2022), 7 waves, 90+ countries, ~443K respondents"
    url = "https://www.worldvaluessurvey.org"
    default_weight = "S017"
    default_psu = ""
    default_strata = ""

    def download(self, years: list[int] | None = None) -> pd.DataFrame:
        if has_cache(WVS_CACHE_KEY):
            df = load_cache(WVS_CACHE_KEY)
            vlabels = _load_meta()
            df.attrs["variable_labels"] = vlabels
        else:
            csv_path = WVS_DATA_DIR / WVS_CSV_NAME
            if not csv_path.exists():
                raise RuntimeError(
                    f"WVS Time Series CSV not found.\n\n"
                    f"Please download the CSV from:\n"
                    f"  {WVS_DOWNLOAD_URL}\n\n"
                    f"Place the CSV file at:\n"
                    f"  {csv_path}\n\n"
                    f"Optionally, also place the variable list Excel file\n"
                    f"(F00003844-...xlsx) in the same directory for variable labels.\n\n"
                    f"Run socdata again and it will be cached automatically."
                )

            from rich.console import Console

            console = Console()
            console.print("[bold]Loading WVS Time Series CSV...[/bold] (this may take a minute)")

            df = pd.read_csv(csv_path, low_memory=False)

            # Strip quotes from headers (CSV uses "quoted" headers)
            df.columns = [c.strip().strip('"').upper() for c in df.columns]

            # Load variable labels from Excel documentation
            vlabels = _load_variable_labels()
            df.attrs["variable_labels"] = vlabels

            # Recode negative missing-data codes to NaN
            df = _recode_missing(df)

            # Cache
            save_cache(WVS_CACHE_KEY, df)
            _save_meta(vlabels)
            console.print(
                f"[green]Cached {len(df):,} respondents × {len(df.columns):,} variables "
                f"as parquet.[/green]"
            )

        # Filter by years if requested (S020 = year of survey)
        if years and "S020" in df.columns:
            df = df[df["S020"].isin(years)].copy()

        return df

    def list_variables(self, df: pd.DataFrame) -> list[str]:
        return list(df.columns)

    def inspect_variables(
        self, df: pd.DataFrame, var_names: list[str]
    ) -> dict[str, VariableInfo]:
        labels = df.attrs.get("variable_labels", {})
        results = {}
        for var in var_names:
            vu = var.upper()
            if vu not in df.columns:
                results[vu] = VariableInfo(name=vu, found=False)
                continue
            col = df[vu]
            n_valid = int(col.notna().sum())
            nunique = col.dropna().nunique()
            vtype = (
                "binary" if nunique == 2
                else "ordinal" if nunique <= 10
                else "continuous" if pd.api.types.is_numeric_dtype(col)
                else "categorical"
            )

            # Wave availability (which of waves 1–7 contain this variable)
            wave_info = ""
            if "S002VS" in df.columns:
                waves_present = sorted(
                    int(w) for w in df.loc[col.notna(), "S002VS"].dropna().unique()
                    if w > 0
                )
                if waves_present:
                    wave_info = f"Waves: {', '.join(str(w) for w in waves_present)}"

            # Build label with wave info appended
            label = labels.get(vu, "")
            if wave_info:
                label = f"{label}  [{wave_info}]" if label else wave_info

            results[vu] = VariableInfo(
                name=vu,
                label=label,
                type=vtype,
                n_valid=n_valid,
                n_missing=int(col.isna().sum()),
                found=True,
            )
        return results

    def system_prompt_appendix(self) -> str:
        return """
**WVS-specific facts:**
- World Values Survey Time Series (1981–2022), 7 waves, 90+ countries
- ~443,000 respondents across all waves
- Weight variable: S017 (within-country weight)
- Equilibrated weight: S018 (normalized to 1000 per country, use for cross-country pooled analysis)
- Country: COUNTRY_ALPHA (3-letter ISO code)
- Wave: S002VS (1–7), Year: S020
- Variable naming convention:
  A = values/attitudes, B = environment/science, C = work/economy
  D = family/marriage, E = politics/governance, F = religion/morality
  G = national identity/immigration, H = security, X = demographics
  S = survey administration, Y = derived indices
- Key demographics: X001 (sex), X003 (age), X025 (education), X047_WVS (income)
- Negative values (e.g., -1, -2, -4, -5) are missing data codes — they have been recoded to NaN
- Common scales: 1–4 (importance), 1–10 (satisfaction/trust), 1–3 (agree/disagree)
"""


register(WVSProvider())
